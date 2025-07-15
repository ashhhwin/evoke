from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json
import time
import gradio as gr
from run_full_pipeline import (
    run_finnhub_pipeline,
    run_eodhd_pipeline,
    run_historical_pipeline,
    load_eps_and_revenue_data, 
    get_latest_daily_date,
    get_ticker_data,
    load_latest_eodhd_merged,
    load_historical_close_prices,
    load_eps_revenue_changes,
    run_pipelines_concurrently,
    load_tickers
)
import datetime
from gradio_calendar import Calendar
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
from compare_eps_revenue import list_available_dates, list_available_periods, compare_eps_revenue
import os
import tempfile
import plotly.express as px
import io
import time
from google.cloud import storage
from datetime import datetime

def get_fixed_periods():
    today = pd.Timestamp.today()

    # Define quarters
    current_q = (today.month - 1) // 3 + 1
    current_y = today.year

    def quarter_str(q, y):
        y_short = str(y)[-2:]  # Convert 2025 → "25"
        return f"Q{q}-{y_short}"

    # Build list
    prev_q = quarter_str(current_q - 1 if current_q > 1 else 4, current_y if current_q > 1 else current_y - 1)
    curr_q = quarter_str(current_q, current_y)
    next_q = quarter_str(current_q + 1 if current_q < 4 else 1, current_y if current_q < 4 else current_y + 1)

    return [
        prev_q,
        curr_q,
        next_q,
        str(current_y),
        str(current_y + 1),
        str(current_y + 2)
    ]

def read_csv_from_gcs(bucket_name, blob_path, max_retries=5, delay=2):
    """
    Read CSV from GCS with retry logic.
    """
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_path)

    print(f"Reading from GCS path: {blob_path}")

    for attempt in range(max_retries):
        if blob.exists():
            content = blob.download_as_text()
            return pd.read_csv(io.StringIO(content))
        else:
            print(f"[Attempt {attempt+1}] File not found: {blob_path}. Retrying in {delay}s...")
            time.sleep(delay)

    raise FileNotFoundError(f"File not found in GCS after {max_retries} attempts: {bucket_name}/{blob_path}")

def get_ticker_list():
    try:
        eps_df, _ = load_eps_and_revenue_data()
        tickers = sorted(eps_df["ticker"].dropna().unique().tolist())
        return tickers
    except Exception as e:
        print(f"Error loading tickers: {e}")
        return [f"Error: {e}"]

def run_finnhub():
    try:
        msg = run_finnhub_pipeline()
        return msg if msg else "Finnhub data download completed."
    except Exception as e:
        return f"Failed: {e}"

def run_eodhd():
    try:
        msg = run_eodhd_pipeline()
        return msg if msg else "OHLCV (EODHD) download completed."
    except Exception as e:
        return f"Failed: {e}"

def run_historical(start, end):
    try:
        start_date = start.date()
        end_date = end.date()
        msg = run_historical_pipeline(start_date, end_date)
        return msg if msg else f"Historical data downloaded from {start_date} to {end_date}"
    except Exception as e:
        return f"Failed: {e}"
    
def get_eps_revenue_changes(ticker_filter=None):
    try:
        df = load_eps_revenue_changes()
        if ticker_filter:
            tickers = [t.strip().upper() for t in ticker_filter.split(",")]
            df = df[df["ticker"].isin(tickers)]
        return df
    except Exception as e:
        return pd.DataFrame({"Error": [f"Error loading EPS/Revenue changes: {e}"]})

def plot_eps_revenue(ticker: str, data_type: str):
    try:
        eps_df, rev_df = load_eps_and_revenue_data()
        eps, rev = get_ticker_data(ticker, eps_df, rev_df)
        # Choose only quarterly or annual columns
        if data_type == "Quarterly":
            cols = [c for c in eps.index if c.startswith("Q")]
        else:
            cols = [c for c in eps.index if not c.startswith("Q") and not c == "api_run_date"]
        # EPS Plot
        eps_plot = go.Figure()
        eps_plot.add_trace(go.Scatter(x=cols, y=eps.loc[cols].values.flatten(), mode='lines+markers', name="EPS"))
        eps_plot.update_layout(title=f"{ticker} EPS ({data_type})", xaxis_title="Period", yaxis_title="EPS")
        # Revenue Plot
        rev_plot = go.Figure()
        rev_plot.add_trace(go.Scatter(x=cols, y=rev.loc[cols].values.flatten(), mode='lines+markers', name="Revenue"))
        rev_plot.update_layout(title=f"{ticker} Revenue ({data_type})", xaxis_title="Period", yaxis_title="Revenue (in millions)")
        return eps_plot, rev_plot
    except Exception as e:
        return go.Figure(layout_title_text=f"Error loading EPS: {e}"), go.Figure(layout_title_text=f"Error loading Revenue: {e}")

def display_latest_ticker_snapshot(ticker: str):
    try:
        row = load_latest_eodhd_merged(ticker)

        display_keys = [
        "Trade_Date", "Symbol", "Company_Name", "Sector", "Industry", "MarketCapitalization", "Beta",
        "P_Open", "P_High", "P_Low", "P_Close", "volume", "Prev_Close (Price)",
        "P_50D_MA", "P_200D_MA", "V_14D_MA", "V_50D_MA", "Options", "hi_250d", "lo_250d",
        "Close_to_Open (% from Prev Day Close)", "Open_Close (%)", "High_Close(%)", "Low_Close(%)",
        "Close_to_Close (%)", "Shares_Out", "Shares_Float", "Short_Ratio", "Short_Percent_Float",
        "Earnings_Date", "Shares_Insiders", "Shares_Institutions"

        ]

        row_data = {k: row[k] for k in display_keys if k in row}

        # Format as row-wise table
        formatted = "\n".join([f"{k} : {v}" for k, v in row_data.items()])
        return formatted

    except Exception as e:
        return f"Error loading latest EODHD data: {e}"
    
def run_both_pipelines():
    try:
        try:
            tickers = load_tickers()
        except Exception as e:
            return f"Failed: {e}"
        run_pipelines_concurrently(tickers)
        return "Both Finnhub and EODHD pipelines completed."
    except Exception as e:
        return f"Failed: {e}"
       
def plot_close_price_history_plotly_without_subplots(ticker: str):
    try:
        df = load_historical_close_prices(ticker)
        # Sort by date
        df = df.sort_values("Trade_Date")
        colors = ["green" if val >= 0 else "red" for val in df["Close_to_Close (%)"]]
        # ----------- First Plot: Close Price -----------
        fig_close = go.Figure()
        fig_close.add_trace(go.Scatter(
            x=df["Trade_Date"],
            y=df["P_Close"],
            mode="lines+markers",
            name="Close Price"
        ))
        fig_close.update_layout(
            title=f"{ticker.upper()} - Close Price",
            xaxis=dict(
                title="Date",
                rangeselector=dict(
                    buttons=list([
                        dict(count=7, label="1W", step="day", stepmode="backward"),
                        dict(count=30, label="1M", step="day", stepmode="backward"),
                        dict(count=6, label="6M", step="month", stepmode="backward"),
                        dict(count=12, label="1Y", step="month", stepmode="backward"),
                        dict(step="all", label="All")
                    ])
                ),
                rangeslider=dict(visible=True),  # Optional: adds a zoom slider
                type="date"
            ),
            yaxis_title="Price",
            height=400
        )
        # ----------- Second Plot: Volume + MA Lines -----------
        fig_vol = go.Figure()
        fig_vol.add_trace(go.Bar(
            x=df["Trade_Date"],
            y=df["volume"],
            name="Volume",
            marker_color=colors
        ))
        if "V_14D_MA" in df.columns:
            fig_vol.add_trace(go.Scatter(
                x=df["Trade_Date"],
                y=df["V_14D_MA"],
                name="14D MA Volume",
                mode="lines",
                line=dict(color="orange", dash="dash")
            ))
        if "V_50D_MA" in df.columns:
            fig_vol.add_trace(go.Scatter(
                x=df["Trade_Date"],
                y=df["V_50D_MA"],
                name="50D MA Volume",
                mode="lines",
                line=dict(color="blue", dash="dot")
            ))
        fig_vol.update_layout(
            title=f"{ticker.upper()} - Volume + 14D/50D MAs",
            xaxis=dict(
                title="Date",
                rangeselector=dict(
                    buttons=list([
                        dict(count=7, label="1W", step="day", stepmode="backward"),
                        dict(count=30, label="1M", step="day", stepmode="backward"),
                        dict(count=6, label="6M", step="month", stepmode="backward"),
                        dict(count=12, label="1Y", step="month", stepmode="backward"),
                        dict(step="all", label="All")
                    ])
                ),
                rangeslider=dict(visible=True),  # Optional: adds a zoom slider
                type="date"
            ),
            yaxis_title="Volume",
            height=400
        )
        return fig_close, fig_vol
    except Exception as e:
        return (
            go.Figure(layout_title_text=f"Error: {e}"),
            go.Figure(layout_title_text=f"No volume data available: {e}")
        )
        
def plot_close_price_history(ticker: str):
    try:
        df = load_historical_close_prices(ticker)
        df = df.sort_values("Trade_Date")
        colors = ["green" if val >= 0 else "red" for val in df["Close_to_Close (%)"]]

        # Create subplot with shared x-axis
        fig = make_subplots(
            rows=2, cols=1,
            shared_xaxes=True,
            vertical_spacing=0.1,
            row_heights=[0.5, 0.5],
            subplot_titles=(f"{ticker.upper()} - Close Price", f"{ticker.upper()} - Volume + 14D/50D MAs")
        )

        # Row 1: Close Price
        fig.add_trace(go.Scatter(
            x=df["Trade_Date"],
            y=df["P_Close"],
            mode="lines+markers",
            name="Close Price"
        ), row=1, col=1)

        # Row 2: Volume Bar
        fig.add_trace(go.Bar(
            x=df["Trade_Date"],
            y=df["volume"],
            name="Volume",
            marker_color=colors
        ), row=2, col=1)

        # Row 2: 14D and 50D MA lines
        if "V_14D_MA" in df.columns:
            fig.add_trace(go.Scatter(
                x=df["Trade_Date"],
                y=df["V_14D_MA"],
                name="14D MA Volume",
                mode="lines",
                line=dict(color="orange", dash="dash")
            ), row=2, col=1)

        if "V_50D_MA" in df.columns:
            fig.add_trace(go.Scatter(
                x=df["Trade_Date"],
                y=df["V_50D_MA"],
                name="50D MA Volume",
                mode="lines",
                line=dict(color="blue", dash="dot")
            ), row=2, col=1)

        # Layout
        fig.update_layout(
            height=800,
            showlegend=True,
            xaxis=dict(
                title="Date"
            ),
            xaxis2=dict(
                title="Date",
                rangeselector=dict(
                    buttons=list([
                        dict(count=7, label="1W", step="day", stepmode="backward"),
                        dict(count=30, label="1M", step="day", stepmode="backward"),
                        dict(count=6, label="6M", step="month", stepmode="backward"),
                        dict(count=12, label="1Y", step="month", stepmode="backward"),
                        dict(step="all", label="All")
                    ]),
                    x=0.5,
                    y=-0.2,
                    xanchor='center',
                    yanchor='top'
                ),
            ),
            yaxis=dict(title="Close Price"),
            yaxis2=dict(title="Volume")
        )

        return fig
    except Exception as e:
        return go.Figure(layout_title_text=f"Error: {e}")
## ashwin changes start here for excel workbook

def load_df(filepath):
    df = pd.read_csv(filepath, keep_default_na=False, na_values=[""," "])
    df = df.rename(columns={
    'ticker': 'Symbol',
    'Company_Name': 'Name',
    'MarketCapitalization': 'Mkt. Cap',
    'Shares_Float': 'Float',
    'Earnings_Date': 'Earnings Date',
    'prev_revenue_millions': 'Prev Revenue',
    'prev_eps': 'Prev EPS',
    'new_revenue_millions': 'New Revenue',
    'new_eps': 'New EPS',
    'revenue_pct_change': '% Revenue',
    'eps_pct_change': '% EPS'
    })
    return df

def transform_to_wrkbook(df):
    base_headers = ["Symbol", "Name", "Type", "Sector", "Industry", "Mkt. Cap", "Float", "Earnings Date"]
    group_data = ["4W Revenue", "4W EPS", "Now Revenue", "Now EPS", "% Revenue", "% EPS"]
    all_headers = base_headers + group_data

    insert_after = ["Earnings Date", "4W EPS", "Now EPS"]
    blank_col_indices = []

    for after_col in insert_after:
        idx = all_headers.index(after_col) + 1
        all_headers.insert(idx, "")
        blank_col_indices.append(idx)

    full_headers = all_headers.copy()

    display_headers = (
        base_headers +
        [""] + ["Revenue", "EPS"] +
        [""] + ["Revenue", "EPS"] +
        [""] + ["Revenue", "EPS"]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Earnings Data"

    center_align = Alignment(horizontal="center", vertical="center")
    bold_yellow_font = Font(bold=True, color="FFFF00")
    light_gray_fill = PatternFill("solid", fgColor="999999")
    dark_gray_fill = PatternFill("solid", fgColor="999999")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    group_col_map = {
        "4 Weeks Ago": ("4W Revenue", "4W EPS"),
        "Present": ("Now Revenue", "Now EPS"),
        "% Change": ("% Revenue", "% EPS")
    }

    def get_column_index(name):
        if name not in full_headers:
            raise ValueError(f"Column '{name}' not found in headers")
        return full_headers.index(name) + 1

    for group_name, (start_col_name, end_col_name) in group_col_map.items():
        col_start = get_column_index(start_col_name)
        col_end = get_column_index(end_col_name)
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        cell = ws.cell(row=1, column=col_start, value=group_name)
        cell.alignment = center_align
        cell.font = bold_yellow_font
        cell.fill = light_gray_fill
        cell.border = thin_border

    for col_idx, header in enumerate(display_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.alignment = center_align
        cell.font = bold_yellow_font
        cell.fill = light_gray_fill
        cell.border = thin_border

    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        data_idx = 0
        for col_idx, header in enumerate(full_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if header == "":
                cell.fill = dark_gray_fill
                cell.value = None
            else:
                value = row[data_idx]
                data_idx += 1

                # Format logic
                cell.value = value
                if header == "Mkt. Cap":
                    cell.number_format = '"$"#,##0'
                elif header == "Float":
                    cell.number_format = '#,##0'
                elif "Revenue" in header and "%" not in header:
                    cell.number_format = '"$"#,##0'
                elif "EPS" in header and "%" not in header:
                    cell.number_format = '"$"#,##0.00'
                elif header in ["% Revenue", "% EPS"]:
                    cell.number_format = '0.0%'
                    if isinstance(cell.value, (int, float)):
                        cell.value = cell.value / 100  # Convert 8.3 to 0.083

            cell.alignment = center_align
            cell.border = thin_border

    # Style any remaining empty merged header cells
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value is None:
            cell.fill = light_gray_fill
        cell.alignment = center_align
        cell.font = bold_yellow_font
        cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    #wb.save("market_data_revisions_eps_revenue_comparison_2025-06-17_to_2025-07-10_for_Q3-25ashwinram.xlsx")
    return wb

def generate_excel_from_comparison_csv(csv_filename: str) -> str:
    import os
    from google.cloud import storage

    bucket = "historical_data_evoke"
    gcs_path = f"market_data/revisions/{csv_filename}"
    local_csv_path = f"/tmp/{csv_filename}"

    client = storage.Client()

    blob = client.bucket(bucket).blob(gcs_path)

    if not blob.exists():
        raise FileNotFoundError(f"GCS file not found: {gcs_path}")

    # Download to local temp file
    blob.download_to_filename(local_csv_path)
    df = load_df(local_csv_path)
    wb = transform_to_wrkbook(df)
    excel_path = local_csv_path.replace(".csv", ".xlsx")
    wb.save(excel_path)
    return excel_path



## ashwin changes end here for excel workbook



def run_comparison(from_date, to_date, period,month=None, selected_caps=None):
    #added month=None
    #added selected_caps = None

    safe_period = str(period).replace(' ', '').replace('/', '').replace('\\', '').replace(':', '')
    output_file = f"eps_revenue_comparison_{from_date}_to_{to_date}_for_{safe_period}.csv"
    compare_eps_revenue(from_date=from_date, to_date=to_date, quarters=[period] if period else None, output_file=output_file, annual=False)
    #print("I AM HERE")
    gcs_path = f'market_data/revisions/{output_file}'
    #print(f"Reading from GCS path: {gcs_path}")
    #time.sleep(20)
    df = read_csv_from_gcs('historical_data_evoke', gcs_path)
    #df = read_csv_from_gcs('historical_data_evoke',f'market_data/revisions/{output_file}')
    
    if df is None or df.empty:
        return None, "No data found for the selected options.", "", "", "", "", ""
    
    df['eps_pct_change'] = pd.to_numeric(df['eps_pct_change'], errors='coerce')
    df['revenue_pct_change'] = pd.to_numeric(df['revenue_pct_change'], errors='coerce')
    df['MarketCapitalization'] = pd.to_numeric(df['MarketCapitalization'], errors='coerce')
    df['Earnings_Date'] = pd.to_datetime(df['Earnings_Date'], errors='coerce')

    if month and month != "All":
        try:
            month_num = datetime.strptime(month, "%B").month
            df = df[df['Earnings_Date'].dt.month == month_num]
        except ValueError:
            pass 
            
    df = df.rename(columns={
    'ticker': 'Symbol',
    'Company_Name': 'Name',
    'MarketCapitalization': 'Mkt. Cap',
    'Shares_Float': 'Float',
    'Earnings_Date': 'Earnings Date',
    'prev_revenue_millions': 'Prev Revenue',
    'prev_eps': 'Prev EPS',
    'new_revenue_millions': 'New Revenue',
    'new_eps': 'New EPS',
    'revenue_pct_change': '% Revenue',
    'eps_pct_change': '% EPS'
    })
    
    def categorize_market_cap(cap):
        if pd.isna(cap): return "Unknown"
        if cap >= 200000: return "Mega Cap"
        elif cap >= 10000: return "Large Cap"
        elif cap >= 2000: return "Mid Cap"
        elif cap >= 250: return "Small Cap"
        elif cap >= 50: return "Micro Cap"
        else: return "Nano Cap"
            

    df['MarketCapCategory'] = df['Mkt. Cap'].apply(categorize_market_cap)
    all_categories = ['Mega Cap', 'Large Cap', 'Mid Cap', 'Small Cap', 'Micro Cap', 'Nano Cap']
    if selected_caps is None or len(selected_caps) == 0:
        selected_caps = all_categories
    df = df[df['MarketCapCategory'].isin(selected_caps)]
    df = df[df['Sector'].notna() & df['Industry'].notna() & df['% EPS'].notna() & df['% Revenue'].notna()]
    df['count'] = 1

    color_scale = [
        [0.0, "yellow"], [0.25, "red"],
        [0.5, "white"], [0.75, "green"], [1.0, "blue"]
    ]

    eps_grouped = df.groupby(['Sector', 'Industry']).agg({'% EPS': 'mean', 'count': 'sum'}).reset_index()
    fig_eps = px.treemap(
        eps_grouped,
        path=['Sector', 'Industry'],
        values='count',
        color='% EPS',
        color_continuous_scale=color_scale,
        color_continuous_midpoint=0,
        custom_data=['% EPS'],
        title='EPS % Change: Sector → Industry'
    )
    fig_eps.update_traces(
        hovertemplate='<b>%{label}</b><br>EPS % Change: %{customdata[0]:.2f}%',
        texttemplate='%{label}<br>%{customdata[0]:.1f}%',
        textposition='middle center'
    )
    eps_plot = fig_eps

    rev_grouped = df.groupby(['Sector', 'Industry']).agg({'% Revenue': 'mean', 'count': 'sum'}).reset_index()
    fig_rev = px.treemap(
        rev_grouped,
        path=['Sector', 'Industry'],
        values='count',
        color='% Revenue',
        color_continuous_scale=color_scale,
        color_continuous_midpoint=0,
        custom_data=['% Revenue'],
        title='Revenue % Change: Sector → Industry'
    )
    fig_rev.update_traces(
        hovertemplate='<b>%{label}</b><br>Revenue % Change: %{customdata[0]:.2f}%',
        texttemplate='%{label}<br>%{customdata[0]:.1f}%',
        textposition='middle center'
    )
    rev_plot = fig_rev
    top_eps_up = df.nlargest(10, '% EPS')[['Symbol', 'Name', 'Mkt. Cap','Prev EPS','New EPS','% EPS']]
    top_eps_down = df.nsmallest(10, '% EPS')[['Symbol', 'Name','Mkt. Cap', 'Prev EPS','New EPS', '% EPS']]
    top_rev_up = df.nlargest(10, '% Revenue')[['Symbol', 'Name','Mkt. Cap','Prev Revenue','New Revenue', '% Revenue']]
    top_rev_down = df.nsmallest(10, '% Revenue')[['Symbol', 'Name','Mkt. Cap', 'Prev Revenue','New Revenue', '% Revenue']]

    '''
    eps_movers_table = (
        "<h4>Top EPS Up</h4>" + top_eps_up.to_html(index=False, escape=False) +
        "<h4>Top EPS Down</h4>" + top_eps_down.to_html(index=False, escape=False)
    )
    rev_movers_table = (
        "<h4>Top Revenue Up</h4>" + top_rev_up.to_html(index=False, escape=False) +
        "<h4>Top Revenue Down</h4>" + top_rev_down.to_html(index=False, escape=False)
    )
    '''
    # Side-by-side EPS movers table
    eps_movers_table = f"<div style='display: flex; gap: 40px; justify-content: space-between;'><div style='flex: 1'><h4>Top EPS Up</h4>{top_eps_up.to_html(index=False, escape=False)}</div><div style='flex: 1'><h4>Top EPS Down</h4>{top_eps_down.to_html(index=False, escape=False)}</div></div>"
    rev_movers_table = f"<div style='display: flex; gap: 40px; justify-content: space-between;'><div style='flex: 1'><h4>Top Revenue Up</h4>{top_rev_up.to_html(index=False, escape=False)}</div><div style='flex: 1'><h4>Top Revenue Down</h4>{top_rev_down.to_html(index=False, escape=False)}</div></div>"
    

    summary = {
        'EPS Up': int((df['% EPS'] > 0).sum()),
        'EPS Down': int((df['% EPS'] < 0).sum()),
        'Revenue Up': int((df['% Revenue'] > 0).sum()),
        'Revenue Down': int((df['% Revenue'] < 0).sum()),
    }
    summary_text = " \
    ".join(f"{k}: {v}" for k, v in summary.items())

    ## ashwin change
    excel_path = generate_excel_from_comparison_csv(output_file)

    return "Comparison and insights complete.", eps_plot, rev_plot, eps_movers_table, rev_movers_table, summary_text, excel_path


# Dates & Periods

dates = list_available_dates()
latest = dates[-1] if dates else None
prior = dates[-2] if len(dates) > 1 else None

def get_periods_for_date(date):
    return list_available_periods(date)

def update_periods(to_date):
    periods = get_periods_for_date(to_date)
    return gr.update(choices=periods, value=periods[0] if periods else None)

def download_csv(file_path):
    return file_path


##ashwin changes start here

def load_news_from_gcs(date_str, ticker, keyword="", bucket_name="historical_data_evoke"):
    """
    Load and format news JSON into uniform HTML cards with optional keyword search.
    """
    import json
    import datetime
    from google.cloud import storage

    blob_path = f"market_data/news/{date_str}/{ticker.upper()}.json"

    try:
        client = storage.Client()
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(blob_path)

        if not blob.exists():
            return f"<div style='color:red; font-weight:bold;'>No news found for {ticker.upper()} on {date_str}.</div>"

        content = blob.download_as_text()
        articles = json.loads(content)

        if not isinstance(articles, list) or not articles:
            return "<div>No valid news entries.</div>"

        keyword = keyword.strip().lower()

        html = """
        <div style='max-height:800px; overflow-y:auto; padding-right:10px;'>
            <div style='display: flex; flex-wrap: wrap; gap: 20px; justify-content: flex-start;'>
        """

        for article in articles:
            headline = article.get("headline", "No headline")
            summary = article.get("summary", "No summary available.")
            source = article.get("source", "Unknown")
            url = article.get("url", "#")
            timestamp = article.get("datetime", None)
            image = article.get("image", "")
            time_str = datetime.datetime.fromtimestamp(timestamp).strftime('%b %d, %Y – %H:%M') if timestamp else "N/A"

            # Keyword filter
            if keyword and keyword not in headline.lower() and keyword not in summary.lower():
                continue

            image_html = f"<img src='{image}' style='width:100%; height:auto; max-height:140px; border-radius:4px; object-fit:cover;'/>" if image else ""

            card = (
                f"<a href='{url}' target='_blank' style='text-decoration: none; flex: 1 1 45%; min-width: 300px; max-width: 48%;'>"
                f"<div style='background-color:#1e1e1e; height: 340px; display: flex; flex-direction: column; justify-content: space-between; "
                f"padding:15px; border-radius:6px; color:white; font-family:Arial, sans-serif; overflow: hidden; "
                f"transition: background-color 0.3s ease; cursor: pointer;' "
                f"onmouseover=\"this.style.backgroundColor='#2c2c2c';\" "
                f"onmouseout=\"this.style.backgroundColor='#1e1e1e';\">"

                f"<div>"
                f"<h3 style='margin-bottom:8px; font-size:1.1em; line-height:1.4; height:48px; overflow:hidden; text-overflow:ellipsis;'>{headline}</h3>"
                f"<p style='margin:4px 0; font-size:0.8em; color:#aaa;'>🕒 {time_str} | 📢 {source}</p>"
                f"<p style='margin-top:10px; font-size:0.95em; line-height:1.5; color:#ddd; height:72px; overflow:hidden; text-overflow:ellipsis;'>{summary}</p>"
                f"</div>"

                f"{image_html}"
                f"</div></a>"
            )

            html += card

        html += "</div></div>"

        return html if html.strip() != "" else "<div>No matching articles found.</div>"

    except Exception as e:
        return f"<div style='color:red;'>Error loading news: {e}</div>"


## ashwin changes end here

## ashwin earnings changes start here

from collections import defaultdict

# Add this function below your existing utilities
def load_earnings_calendar_json(from_date, to_date, bucket_name="historical_data_evoke"):
    import datetime, json
    from google.cloud import storage

    from_dt = from_date.date() if isinstance(from_date, datetime.datetime) else datetime.datetime.strptime(str(from_date), "%Y-%m-%d").date()
    to_dt = to_date.date() if isinstance(to_date, datetime.datetime) else datetime.datetime.strptime(str(to_date), "%Y-%m-%d").date()

    client = storage.Client()
    bucket = client.bucket(bucket_name)

    prefix = "market_data/earnings_calendar/2025-01-01_to_2025-12-01"
    blobs = client.list_blobs(bucket_name, prefix=prefix)

    all_entries = []

    for blob in blobs:
        if not blob.name.endswith(".json"):
            continue
        try:
            content = blob.download_as_text()
            parsed = json.loads(content)
            for entry in parsed.get("earningsCalendar", []):
                try:
                    entry_date = datetime.datetime.strptime(entry["date"], "%Y-%m-%d").date()
                except Exception as e:
                    # print(f"❌ Bad date format in {blob.name}: {entry.get('date')}")
                    continue

                if from_dt <= entry_date <= to_dt:
                    # print(f"✅ Matched: {entry['symbol']} for {entry['date']} from {blob.name}")
                    all_entries.append(entry)
        except Exception as e:
            # print(f"Error reading {blob.name}: {e}")
            continue

    return all_entries


def render_earnings_calendar(entries, ticker_filter=""):
    import datetime
    from collections import defaultdict

    grouped = defaultdict(list)
    for entry in entries:
        if ticker_filter and ticker_filter.lower() not in entry["symbol"].lower():
            continue
        grouped[entry["date"]].append(entry)

    grouped = dict(sorted(grouped.items(), key=lambda x: x[0]))
    html = """
    <div style='display: flex; flex-direction: row; overflow-x: auto; gap: 20px; padding: 10px 0;'>
    """

    today = datetime.date.today()

    for date_str, items in grouped.items():
        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        date_label = date_obj.strftime("%a, %b %d")
        ticker_count = len(items)
        header = f"{date_label} <span style='color:#888;'>({ticker_count} Earnings)</span>"
        highlight = "border: 2px solid #00ff9d;" if date_obj == today else ""

        block = f"""
        <div style='min-width: 220px; max-width: 240px; background:#1e1e1e; padding:12px 14px; border-radius:8px; color:#eee; font-family:Inter, sans-serif; box-shadow: 0 0 4px #00000033; {highlight}'>
            <h4 style='color:#00ff9d; font-weight:bold; border-bottom:1px solid #333; padding-bottom:6px; margin-bottom:10px;'>{header}</h4>
        """

        for e in items:
            symbol = e.get("symbol", "—")
            hour = e.get("hour", "tbd").upper()
            if hour == "AMC":
                timing = "After Market"
            elif hour == "BMO":
                timing = "Before Market"
            else:
                timing = "TBD"

            block += f"""
            <div style='padding:6px 0; border-bottom:1px dashed #444;'>
                <div style='font-weight:bold;'>{symbol}</div>
                <div style='font-size:0.85em; color:#aaa;'>{timing}</div>
            </div>
            """

        block += "</div>"
        html += block

    html += "</div>"
    return html if grouped else "<div style='color:red;'>No earnings found for selected range.</div>"

## ashwin earnings changes end here

with gr.Blocks(theme=gr.themes.Soft()) as app:
    gr.Markdown("<h1 style='text-align:center; color:#00ff9d;'>📈 Market Data Dashboard</h1>")

    with gr.Tab("Download Daily Data"):
        gr.Markdown("## Select Data to Download")
        with gr.Row():
            finnhub_btn = gr.Button("Run EPS & Revenue (Finnhub)")
            eodhd_btn = gr.Button("Run OHLCV (EODHD)")
            both_btn = gr.Button("Run Both")
        status_output = gr.Textbox(label="Status", lines=2)
        with gr.Row():
            last_run_label = gr.Textbox(
                value=f"Latest Run Date: {get_latest_daily_date()}",
                label="Last Run (Daily Folder)",
                interactive=False
            )
        finnhub_btn.click(fn=run_finnhub, outputs=status_output)
        eodhd_btn.click(fn=run_eodhd, outputs=status_output)
        both_btn.click(fn=run_both_pipelines, outputs=status_output)

    with gr.Tab("Download Historical Data"):
        gr.Markdown("## Historical OHLCV Download")
        start_date = Calendar(type="date", label="Start Date")
        end_date = Calendar(type="date", label="End Date")
        run_btn = gr.Button("Run Historical Download")
        hist_status = gr.Textbox(label="Status", lines=2)

        run_btn.click(fn=run_historical, inputs=[start_date, end_date], outputs=hist_status)


    with gr.Tab("Ticker Insights"):
        gr.Markdown("## View EPS & Revenue by Ticker")

        with gr.Row():
            ticker_dropdown = gr.Dropdown(
                label="Select Ticker",
                choices=get_ticker_list(),
                interactive=True,
                filterable=True,
                allow_custom_value=False,
                show_label=True,
                scale=2
            )
            data_type = gr.Radio(
                choices=["Quarterly", "Annual"],
                value="Quarterly",
                label="Data Type"
            )

        # EPS and Revenue side-by-side
        with gr.Row():
            eps_plot = gr.Plot(label="EPS Plot")
            rev_plot = gr.Plot(label="Revenue Plot")

        gr.Markdown("## View Close Price and Volume by Ticker")
        # Close price chart below
        #close_plot = gr.Plot(label="Close Price History")
        #volume_plot = gr.Plot(label="Volume ")
        close_vol_plot = gr.Plot(label = " Close Price and Volume")
        # Ticker snapshot info at the bottom
        ticker_info = gr.Textbox(label="Latest Ticker Snapshot", lines=30, interactive=False)

        def update_all(ticker, mode):
            try:
                eps_fig, rev_fig = plot_eps_revenue(ticker, mode)
            except Exception as e:
                eps_fig = go.Figure(layout_title_text=f"Error loading EPS: {e}")
                rev_fig = go.Figure(layout_title_text=f"Error loading Revenue: {e}")
            #try:
            #    price_fig, vol_fig = plot_close_price_history(ticker)
            #except Exception as e:
             #   price_fig = go.Figure(layout_title_text=f"Error loading price: {e}")
              #  vol_fig = go.Figure(layout_title_text=f"Error loading volume: {e}")
            try:
                close_vol_fig = plot_close_price_history(ticker)
            except Exception as e:
                close_vol_fig = go.Figure(layout_title_text=f"Error loading price/volume: {e}")
            try:
                snapshot = display_latest_ticker_snapshot(ticker)
            except Exception as e:
                snapshot = f"Error loading ticker snapshot: {e}"
            #return eps_fig, rev_fig, price_fig, vol_fig, snapshot
            return eps_fig, rev_fig, close_vol_fig, snapshot
    
        # Update all elements when ticker or type changes
        ticker_dropdown.change(
            fn=update_all,
            inputs=[ticker_dropdown, data_type],
            #outputs=[eps_plot, rev_plot, close_plot, volume_plot, ticker_info]
            outputs=[eps_plot, rev_plot, close_vol_plot, ticker_info]
        )

        data_type.change(
            fn=update_all,
            inputs=[ticker_dropdown, data_type],
            #outputs=[eps_plot, rev_plot, close_plot, volume_plot, ticker_info]
            outputs=[eps_plot, rev_plot, close_vol_plot, ticker_info]
        )

    with gr.Tab("EPS & Revenue Revisions"):
        gr.Markdown("### Compare EPS & Revenue Estimates")
        with gr.Row():
            from_date = gr.Dropdown(label="From Date", choices=dates, value=prior)
            to_date = gr.Dropdown(label="To Date", choices=dates, value=latest)
        #period = gr.Dropdown(label="Period", choices=get_periods_for_date(latest), value=get_periods_for_date(latest))
        fixed_periods = get_fixed_periods()
        period = gr.Dropdown(label="Period", choices=fixed_periods, value=fixed_periods[1])  # Default: current quarter
        months = [datetime(2000, m, 1).strftime("%B") for m in range(1, 13)]
        month_dropdown = gr.Dropdown(label="Filter by Earnings Month", choices=['ALL']+months, value="ALL")

        market_caps = ['Mega Cap', 'Large Cap', 'Mid Cap', 'Small Cap', 'Micro Cap', 'Nano Cap']
        market_cap_dropdown = gr.CheckboxGroup(
            choices=market_caps,
            value=market_caps,  # All selected by default
            label="Filter by Market Cap Category"
        )
        
        run_comparison_btn = gr.Button("Run Comparison")
        #output_file = gr.File(label="Download CSV", visible=False)
        status = gr.Textbox(label="Status", interactive=False)

        gr.Markdown("### EPS % Change Treemap")
        eps_treemap_plot = gr.Plot(label="EPS Treemap")
        gr.Markdown("### Revenue % Change Treemap")
        rev_treemap_plot = gr.Plot(label="Revenue Treemap")

        gr.Markdown("### Top EPS Movers")
        eps_movers_table = gr.HTML()
        #gr.HTML(eps_movers_table)
        
        gr.Markdown("### Top Revenue Movers")
        rev_movers_table = gr.HTML()
        #gr.HTML(rev_movers_table)

        gr.Markdown("### Summary of Revisions")
        summary_box = gr.Textbox(lines=8, interactive=False)

        to_date.change(fn=update_periods, inputs=[to_date], outputs=period)

        excel_download = gr.File(label="Download Excel Workbook", interactive=False)

        run_comparison_btn.click(
            fn=run_comparison,
            inputs=[from_date, to_date, period, month_dropdown,market_cap_dropdown],
            outputs=[ status, eps_treemap_plot, rev_treemap_plot, eps_movers_table, rev_movers_table, summary_box, excel_download]
        )

## ashwin changes start here

    with gr.Tab("Market News by Ticker"):
        gr.Markdown("## 📰 Market News by Ticker and Date")
    
        with gr.Row():
            news_date = gr.Dropdown(label="Select Date", choices=dates, value=latest)
            news_ticker = gr.Dropdown(label="Select Ticker", choices=get_ticker_list(), value=None)
    
        search_input = gr.Textbox(label="Search Headline or Summary", placeholder="Type to filter...", lines=1)
    
        load_news_btn = gr.Button("Load News")
        news_output = gr.HTML(label="News Feed")
    
        load_news_btn.click(
            fn=load_news_from_gcs,
            inputs=[news_date, news_ticker, search_input],
            outputs=news_output
        )

    ##ashwin changes end here
    with gr.Tab("Earnings Calendar"):
        gr.Markdown("## 📆 Upcoming Earnings Calendar")
    
        from datetime import date, timedelta
        default_from = (date.today() - timedelta(days=2)).strftime("%Y-%m-%d")
        default_to = (date.today() + timedelta(days=5)).strftime("%Y-%m-%d")
    
        with gr.Row():
            from_cal = Calendar(label="From Date", value=default_from)
            to_cal = Calendar(label="To Date", value=default_to)
            ticker_input = gr.Textbox(label="Search Ticker (optional)", placeholder="e.g. AAPL, TSLA")
    
        load_btn = gr.Button("Load Calendar")
        status_box = gr.Textbox(label="", interactive=False, visible=True, lines=1)
        calendar_output = gr.HTML()
    
        def update_calendar(from_date, to_date, ticker_filter):
            import datetime
            entries = load_earnings_calendar_json(from_date, to_date)
            filtered = [e for e in entries if not ticker_filter or ticker_filter.lower() in e['symbol'].lower()]
            summary = f"✅ Loaded {len(filtered)} earnings from {len(entries)} entries ({from_date} to {to_date})"
            # print(f"👀 Showing {len(filtered)} filtered entries")
            return summary, render_earnings_calendar(filtered, "")
    
        load_btn.click(
            fn=update_calendar,
            inputs=[from_cal, to_cal, ticker_input],
            outputs=[status_box, calendar_output]
        )



## app.launch(server_name="0.0.0.0", server_port=7869)
